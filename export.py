"""
Export Module

Provides functions to export expense data to CSV or Excel files.
"""

import csv
from typing import List, Tuple, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from database import Database


class Export:
    def __init__(self, db: Database) -> None:
        """
        Initialize the Export module with a database instance.
        """
        self.db = db

    def to_csv(self, filename: str) -> bool:
        """
        Export expense data to a CSV file.
        """
        try:
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                # Write header row
                writer.writerow(["ID", "Date", "Amount (₹)", "Category", "Description"])
                # Write each expense row after formatting the amount
                for expense in self.db.get_expenses():
                    expense_list = list(expense)
                    expense_list[2] = self.convert_to_inr(expense_list[2])
                    writer.writerow(expense_list)
            return True
        except Exception as e:
            print("CSV export error:", e)
            return False

    def to_excel(self, filename: str) -> bool:
        """
        Export expense data to an Excel file with additional summary sheets.
        """
        try:
            wb = Workbook()
            # Main Expenses Sheet
            ws = wb.active
            ws.title = "Expenses"
            headers = ["ID", "Date", "Amount (₹)", "Category", "Description"]
            ws.append(headers)
            for row in self.db.get_expenses():
                formatted_row = list(row)
                formatted_row[2] = self.convert_to_inr(formatted_row[2])
                ws.append(formatted_row)

            # Create a DataFrame for summary sheets
            data = pd.DataFrame(
                self.db.get_expenses(), 
                columns=["id", "date", "amount", "category", "description"]
            )
            if not data.empty:
                data['date'] = pd.to_datetime(data['date'], errors='coerce')
            else:
                data = pd.DataFrame(columns=["id", "date", "amount", "category", "description"])

            # Monthly Summary Sheet
            monthly = data.set_index('date').resample("M").sum().reset_index()
            monthly['date'] = monthly['date'].dt.strftime('%Y-%m')
            ws_monthly = wb.create_sheet("Monthly Summary")
            for r in dataframe_to_rows(monthly, index=False, header=True):
                ws_monthly.append(r)

            # Weekly Summary Sheet
            weekly = data.set_index('date').resample("W-Mon").sum().reset_index()
            weekly['date'] = weekly['date'].dt.strftime('%Y-%m-%d')
            ws_weekly = wb.create_sheet("Weekly Summary")
            for r in dataframe_to_rows(weekly, index=False, header=True):
                ws_weekly.append(r)

            # Yearly Summary Sheet
            yearly = data.set_index('date').resample("Y").sum().reset_index()
            yearly['date'] = yearly['date'].dt.year
            ws_yearly = wb.create_sheet("Yearly Summary")
            for r in dataframe_to_rows(yearly, index=False, header=True):
                ws_yearly.append(r)

            self._format_workbook(wb)
            wb.save(filename)
            return True
        except Exception as e:
            print("Excel export error:", e)
            return False

    def _format_workbook(self, wb: Workbook) -> None:
        """
        Apply formatting and styling to all worksheets in the workbook.
        """
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        alignment = Alignment(horizontal="center")

        for sheet in wb.worksheets:
            # Style header row
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
            # Adjust column widths based on maximum cell length
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = (max_length + 2) * 1.2

            # Format the Amount column in the "Expenses" sheet as currency
            if sheet.title == "Expenses":
                for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell.number_format = '₹#,##0.00'

    def convert_to_inr(self, amount: float) -> str:
        """
        Format the given amount as a string in INR currency format.
        """
        return f"₹{amount:,.2f}"
