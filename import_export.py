"""
Import/Export Module

Handles importing and exporting expense data from CSV/Excel files,
and backing up/restoring the SQLite database.
"""

import csv
import os
import shutil
from typing import List, Any, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from database import Database

# Required columns for import processing
REQUIRED_COLUMNS = ["date", "amount", "category", "description"]


class Export:
    """Handles exporting expense data to CSV and Excel files."""

    def __init__(self, db: Database) -> None:
        """
        Initialize the Export module with a database instance.
        """
        self.db = db

    def to_csv(self, filename: str) -> bool:
        """
        Export expense data to a CSV file.
        """
        try:
            expenses: List[Tuple[Any, ...]] = self.db.get_expenses()
            if not expenses:
                print("No expenses to export.")
                return False

            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Date", "Amount (₹)", "Category", "Description"])
                for expense in expenses:
                    # Convert amount to INR format before writing
                    expense_list = list(expense)
                    expense_list[2] = self.convert_to_inr(expense_list[2])
                    writer.writerow(expense_list)

            print(f"CSV export successful: {filename}")
            return True
        except Exception as e:
            print("CSV export error:", e)
            return False

    def to_excel(self, filename: str) -> bool:
        """
        Export expense data to an Excel file with summary sheets.
        """
        try:
            expenses: List[Tuple[Any, ...]] = self.db.get_expenses()
            if not expenses:
                print("No expenses to export.")
                return False

            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"
            headers = ["ID", "Date", "Amount (₹)", "Category", "Description"]
            ws.append(headers)
            for row in expenses:
                formatted_row = list(row)
                formatted_row[2] = self.convert_to_inr(formatted_row[2])
                ws.append(formatted_row)

            # Prepare a DataFrame for generating summary sheets
            df = pd.DataFrame(expenses, columns=["id", "date", "amount", "category", "description"])
            if df.empty:
                df = pd.DataFrame(columns=["id", "date", "amount", "category", "description"])
            else:
                df['date'] = pd.to_datetime(df['date'], errors='coerce')

            # Create summary sheets for monthly, weekly, and yearly data
            self._add_summary_sheets(wb, df, "M", "Monthly Summary")
            self._add_summary_sheets(wb, df, "W-Mon", "Weekly Summary")
            self._add_summary_sheets(wb, df, "Y", "Yearly Summary")

            self._format_workbook(wb)
            wb.save(filename)
            print(f"Excel export successful: {filename}")
            return True
        except Exception as e:
            print("Excel export error:", e)
            return False

    def _add_summary_sheets(self, wb: Workbook, df: pd.DataFrame, freq: str, sheet_name: str) -> None:
        """
        Add a time-based summary sheet to the workbook.
        """
        summary = df.set_index('date').resample(freq).sum().reset_index()
        if freq == "Y":
            summary['date'] = summary['date'].dt.year
        else:
            summary['date'] = summary['date'].dt.strftime('%Y-%m-%d')

        ws_summary = wb.create_sheet(sheet_name)
        for row in dataframe_to_rows(summary, index=False, header=True):
            ws_summary.append(row)

    def _format_workbook(self, wb: Workbook) -> None:
        """
        Apply styling and formatting to all worksheets in the workbook.
        """
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
        alignment = Alignment(horizontal="center")

        for sheet in wb.worksheets:
            # Format header row
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
            # Adjust column widths
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = (max_length + 2) * 1.2
            # Format the Amount column in the "Expenses" sheet as currency
            if sheet.title == "Expenses":
                for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell.number_format = '₹#,##0.00'

    def convert_to_inr(self, amount: float) -> str:
        """
        Convert the given amount to a string formatted in INR.
        """
        return f"₹{amount:,.2f}"


class Import:
    """Handles importing expense data from CSV and Excel files."""

    def __init__(self, db: Database) -> None:
        """
        Initialize the Import module with a database instance.
        """
        self.db = db

    def from_csv(self, filename: str) -> bool:
        """
        Import expense data from a CSV file.
        """
        try:
            df = pd.read_csv(filename)
            return self._process_import(df)
        except Exception as e:
            print("CSV import error:", e)
            return False

    def from_excel(self, filename: str) -> bool:
        """
        Import expense data from an Excel file.
        """
        try:
            df = pd.read_excel(filename, sheet_name="Expenses")
            return self._process_import(df)
        except Exception as e:
            print("Excel import error:", e)
            return False

    def _process_import(self, df: pd.DataFrame) -> bool:
        """
        Process and validate the imported DataFrame.
        """
        # Normalize column names
        df.columns = [col.strip().lower() for col in df.columns]
        missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
        if missing:
            print(f"Import error: Missing required columns {missing}")
            return False

        # Process required columns
        df = df[REQUIRED_COLUMNS].copy()
        df.columns = ["Date", "Amount", "Category", "Description"]
        df["Amount"] = df["Amount"].apply(self._convert_to_inr)

        for _, row in df.iterrows():
            self.db.insert_expense(row["Date"], row["Amount"], row["Category"], row["Description"])

        print("Import successful!")
        return True

    def _convert_to_inr(self, amount: Any) -> float:
        """
        Convert the given amount to a float representing INR.
        """
        try:
            return round(float(amount), 2)
        except Exception as e:
            print(f"Amount conversion error: {e}, received: {amount}")
            return 0.0


class BackupRestore:
    """Handles database backup and restore operations."""

    @staticmethod
    def backup_database(db_filename: str, backup_filename: str) -> bool:
        """
        Create a backup of the database file.
        """
        try:
            shutil.copy(db_filename, backup_filename)
            print(f"Backup successful: {backup_filename}")
            return True
        except Exception as e:
            print("Backup error:", e)
            return False

    @staticmethod
    def restore_database(backup_filename: str, db_filename: str) -> bool:
        """
        Restore the database from a backup file.
        """
        try:
            shutil.copy(backup_filename, db_filename)
            print(f"Database restored from {backup_filename}")
            return True
        except Exception as e:
            print("Restore error:", e)
            return False
